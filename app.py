import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# 1 - Entrar na planilha e extrair o CPF do cliente.
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

# 2 - Entro no site https://consultcpf-devaprender.netlify.app/ o CPF da planilha para pesquisar o status do pagamento daquele cliente
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha

    sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    
    # 3 - Verificar se está "em dia" ou "atrasado"
    botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)
   
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    if status.text == 'em dia':
        # 4 - Se estiver "em dia", pegar a data do pagamento e o método de pagamento
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        data_pagamento_limpo = data_pagamento.text.split()[3]
        
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]
        
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia' , data_pagamento_limpo , metodo_pagamento_limpo])
        
        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        # 5 - caso contrário(se estiver atrasado), colocar o status como pendente
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilha_fechamento.save('planilha fechamento.xlsx')